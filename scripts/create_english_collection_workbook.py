from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "DT_Lab_Website_Information_Collection_Form.xlsx"

NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
GREY = "666666"
THIN = Side(style="thin", color="D9E1F2")


def setup(ws, title, guidance, headers, widths):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = guidance
    ws["A2"].font = Font(size=10, italic=True, color=GREY)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34
    for index, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(3, index, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[3].height = 36
    ws.auto_filter.ref = f"A3:{ws.cell(3, len(headers)).coordinate}"


def add_rows(ws, rows):
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        ws.row_dimensions[row[0].row].height = 38


def status_dropdown(ws, column, end=300):
    dv = DataValidation(
        type="list",
        formula1='"Please review,Confirmed,Updated,New entry,Do not publish,Not applicable"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"{column}4:{column}{end}")


def dropdown(ws, formula, cell_range):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def highlight_input(ws, cell_range):
    first = cell_range.split(":")[0]
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f"LEN({first})=0"], fill=PatternFill("solid", fgColor=YELLOW)),
    )


wb = Workbook()
start = wb.active
start.title = "Start Here"
start.sheet_view.showGridLines = False
start.column_dimensions["A"].width = 24
start.column_dimensions["B"].width = 96
start.merge_cells("A1:B1")
start["A1"] = "DT Lab Website Information Form"
start["A1"].font = Font(size=22, bold=True, color=WHITE)
start["A1"].fill = PatternFill("solid", fgColor=NAVY)
start["A1"].alignment = Alignment(vertical="center")
start.row_dimensions[1].height = 40

welcome = [
    ("Thank you!", "This form helps us keep the DT Lab website accurate and up to date. Most existing information is already included—please review it, make any corrections, and add anything that is missing."),
    ("How to complete it", "Open the relevant tab(s), review each pre-filled row, and type corrections or new information directly into the indicated fields. Add more rows whenever needed."),
    ("Yellow cells", "These are the main fields that need your attention. If a field does not apply, you may leave it blank or select “Not applicable”."),
    ("Status", "Please choose a status for each row: Confirmed, Updated, New entry, Do not publish, or Not applicable. Leave “Please review” if someone else still needs to check it."),
    ("Your name", "Add your name in the “Completed / confirmed by” column so we know whom to contact if clarification is needed."),
    ("Links", "Please use complete URLs, such as https://example.com. For email addresses, enter name@example.edu."),
    ("Photos", "For member profiles, please provide a clear JPG or PNG portrait (square or portrait orientation is ideal). Enter the file name in the form and submit the image with this workbook."),
    ("Dates", "Please use YYYY-MM-DD, for example 2026-07-31. If only the month is known, use the first day of that month and explain this in Notes."),
    ("Privacy", "Only include personal details, email addresses, photos, and links that the person has agreed may be published on the public lab website."),
    ("Need help?", "If you are unsure about a field, leave it blank and add a short note. We can follow up before publishing."),
]
for row_number, (heading, text) in enumerate(welcome, 3):
    start.cell(row_number, 1, heading)
    start.cell(row_number, 1).font = Font(bold=True, color=NAVY)
    start.cell(row_number, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    start.cell(row_number, 2, text)
    for cell in start[row_number]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=THIN)
    start.row_dimensions[row_number].height = 42

start.merge_cells("A15:B15")
start["A15"] = "Quick checklist before returning the form"
start["A15"].font = Font(size=14, bold=True, color=WHITE)
start["A15"].fill = PatternFill("solid", fgColor=BLUE)
checks = [
    "☐ Lab name, description, affiliation, and contact details have been reviewed.",
    "☐ Each member has approved the details and photo intended for publication.",
    "☐ Research themes are accurate and written for a general academic audience.",
    "☐ Publication titles, author order, venues, years, and links have been checked.",
    "☐ News dates, descriptions, and links have been checked.",
]
for row_number, text in enumerate(checks, 16):
    start.cell(row_number, 2, text)
    start.cell(row_number, 2).alignment = Alignment(wrap_text=True)
    start.row_dimensions[row_number].height = 27


lab = wb.create_sheet("Lab Details")
headers = ["Section", "Information needed", "Current website content", "Your confirmed or updated content", "Status", "Completed / confirmed by", "Notes"]
setup(lab, "Lab Details", "Please review the current information and enter the final wording we should publish. Fields marked “Optional” may be left blank.", headers, [17, 29, 52, 55, 20, 25, 40])
rows = [
    ["Identity", "Official lab name (English) — Required", "Data Transpose Lab", "", "Please review", "", "Please confirm whether the preferred name is Data Transpose Lab, DT Lab, or another name."],
    ["Identity", "Official lab name (other language) — Optional", "", "", "Please review", "", ""],
    ["Identity", "Preferred short name — Required", "DT Lab", "", "Please review", "", ""],
    ["Identity", "Short website description (English) — Required", "AI, big data, and cloud-computing research at Adelaide University", "", "Please review", "", "One or two clear sentences is ideal."],
    ["Identity", "Short website description (other language) — Optional", "", "", "Please review", "", ""],
    ["Affiliation", "Official university name — Required", "Adelaide University", "", "Please review", "", "Please use the university's current official English name."],
    ["Affiliation", "School / Faculty — Required", "School of Computer Science and Information Technology", "", "Please review", "", ""],
    ["Affiliation", "Institute / Centre — Optional", "Australian Institute for Machine Learning", "", "Please review", "", ""],
    ["Contact", "Public lab email — Optional", "", "", "Please review", "", "Enter “Not applicable” if there is no shared lab email."],
    ["Contact", "Public website URL — Required", "", "", "Please review", "", "This may be added once the final domain or GitHub Pages address is known."],
    ["Contact", "Full postal address — Required", "The University of Adelaide, North Terrace, Adelaide, South Australia", "", "Please review", "", "Please include building, room, and postcode if appropriate."],
    ["Contact", "Office / room — Optional", "", "", "Please review", "", ""],
    ["Contact", "Public phone number — Optional", "", "", "Please review", "", ""],
    ["Online profiles", "GitHub organisation — Optional", "", "", "Please review", "", ""],
    ["Online profiles", "LinkedIn or other social profile — Optional", "", "", "Please review", "", ""],
]
add_rows(lab, rows)
status_dropdown(lab, "E")
highlight_input(lab, "D4:D200")


members = wb.create_sheet("Lab Members")
headers = [
    "Member category — Required", "Full name for the website — Required", "Preferred title / lab role — Required",
    "Affiliation — Optional", "Research interests — Required", "Short biography — Optional", "Public email — Optional",
    "Personal / university profile — Optional", "Google Scholar — Optional", "GitHub — Optional", "LinkedIn — Optional",
    "Photo file name — Recommended", "Location — Optional", "Status", "Completed / confirmed by", "Notes",
]
setup(members, "Lab Members", "One person per row. Please only include details that the member has agreed may be published. Add rows at the bottom as needed.", headers, [24, 31, 31, 44, 58, 58, 30, 42, 38, 34, 34, 29, 24, 20, 25, 40])
rows = [
    ["Faculty", "Dr. Tony Weitong Chen", "Co-Director and Joint Principal Leader", "School of Computer Science and Information Technology; Australian Institute for Machine Learning", "Foundation model alignment, next-generation neural architectures, trustworthy machine learning, multimodal health data analytics, and machine unlearning.", "Senior Lecturer and ARC Externally-Funded Senior Research Fellow at Adelaide University. Eligible to supervise Master's and PhD students.", "weitong.chen@adelaide.edu.au", "https://researchers.adelaide.edu.au/profile/weitong.chen", "", "", "", "weitong-chen.jpg", "Adelaide, Australia", "Please review", "", "Please check the title, role, biography, and preferred public name."],
    ["Faculty", "A/Prof. Wei Emma Zhang", "Co-Director and Joint Principal Leader", "School of Computer Science and Information Technology; Australian Institute for Machine Learning", "Natural language processing, text mining, information retrieval, multimodal generation and evaluation, federated learning in NLP, and Artificial Intelligence of Things.", "Associate Professor and ARC Externally-Funded Senior Research Fellow at Adelaide University. Eligible to supervise Master's and PhD students.", "wei.e.zhang@adelaide.edu.au", "https://researchers.adelaide.edu.au/profile/wei.e.zhang", "", "", "", "wei-emma-zhang.jpg", "Adelaide, Australia", "Please review", "", "Please check the title, role, biography, and preferred public name."],
    ["PhD Student", "Wenhao Liang", "PhD Candidate", "Australian Institute for Machine Learning; SA Pathology; The University of Adelaide", "Trust, safety, security, privacy, uncertainty estimation, model calibration, and model optimization.", "", "", "http://eaglewhliang.github.io/", "https://scholar.google.com/citations?user=3f4N-oQAAAAJ&hl=en", "https://github.com/EagleAdelaide", "", "wenhao-liang.jpg", "Adelaide, Australia", "Please review", "", "Please add a biography and public email if desired."],
]
rows.extend([["", "", "", "", "", "", "", "", "", "", "", "", "", "New entry", "", ""] for _ in range(15)])
add_rows(members, rows)
status_dropdown(members, "N")
dropdown(members, '"Faculty,Postdoctoral Researcher,PhD Student,Master Student,Undergraduate Student,Research Staff,Visitor,Alumni,Other"', "A4:A300")
highlight_input(members, "B4:B300")


research = wb.create_sheet("Research Themes")
headers = ["Display order", "Theme title — Required", "Current website description", "Your confirmed or updated description — Required", "Keywords — Optional", "Feature on home page?", "Status", "Completed / confirmed by", "Notes"]
setup(research, "Research Themes", "Please use clear, accessible English. A description of approximately 25–60 words per theme works well on the website.", headers, [14, 34, 66, 68, 48, 22, 20, 25, 40])
rows = [
    [1, "Trustworthy AI", "Model calibration, uncertainty estimation, fairness, adversarial robustness, privacy, interpretability, and explainability.", "", "", "Yes", "Please review", "", ""],
    [2, "AI and Data Mining", "NLP, text mining, causal reasoning, multimodal information fusion, and knowledge discovery from imperfect data.", "", "", "Yes", "Please review", "", ""],
    [3, "Model Optimization", "Efficient training and inference, hyperparameter and optimization methods, transfer learning, meta-learning, and hardware-aware AI.", "", "", "Yes", "Please review", "", ""],
    [4, "Big Data and Cloud Computing", "Big data, distributed systems, and cloud computing.", "", "", "Please review", "Please review", "", "Please confirm whether this should be featured on the home page."],
]
rows.extend([[i, "", "", "", "", "", "New entry", "", ""] for i in range(5, 10)])
add_rows(research, rows)
status_dropdown(research, "G")
dropdown(research, '"Yes,No,Please review"', "F4:F300")
highlight_input(research, "D4:D300")


pubs = wb.create_sheet("Publications")
headers = ["Title — Required", "Authors in published order — Required", "Venue — Required", "Year — Required", "Type — Optional", "Short summary / highlight — Optional", "DOI URL", "arXiv URL", "PDF URL", "Code URL", "Project URL", "Status", "Completed / confirmed by", "Notes"]
setup(pubs, "Publications", "One publication per row. Please copy the title, author order, venue, and year from the official publication record whenever possible.", headers, [58, 68, 38, 16, 20, 62, 38, 38, 38, 38, 38, 20, 25, 42])
rows = [
    ["Calibrating on Kolmogorov-Arnold Networks", "Wenhao Liang, Wei Emma Zhang, Lin Yue, Miao Xu, Olaf Maennel, Weitong Chen", "CIKM 2025, Seoul", 2025, "", "", "", "https://arxiv.org/abs/2503.01195", "", "", "", "Please review", "", "Please verify the citation and add any available links."],
    ["We Care Each Pixel: Calibrating on Medical Segmentation Models through Signed Distance", "Wenhao Liang, Wei Emma Zhang, Lin Yue, Miao Xu, Olaf Maennel, Weitong Chen", "CIKM 2025, Seoul", 2025, "", "", "", "https://arxiv.org/abs/2503.05107", "", "", "", "Please review", "", "Please verify the citation and add any available links."],
    ["TraffiX-MoE: A Traffic-Aware Neural VRP Solver", "Wenhao Liang, Lin Yue, Wei Emma Zhang, Joy Rathjen, Peter O'Loughlin, Weitong Chen", "ADMA 2025, Kyoto, Industry Track", 2025, "", "", "", "", "", "", "", "Please review", "", "Please add a paper, DOI, or project link."],
    ["Enhancing Financial Market Predictions: Causality-Driven Feature Selection", "Wenhao Liang, Zhengyang Li, Weitong Chen", "ADMA 2024", 2024, "", "Integrates financial news and stock-market data across 197 countries with LSTM models to improve market-prediction accuracy.", "", "", "", "", "", "Please review", "", "Please verify the summary and add available links."],
    ["Correlation Analysis of Adversarial Attack in Time Series Classification", "Zhengyang Li, Wenhao Liang, Chang Dong, Weitong Chen, Dong Huang", "ADMA 2024", 2024, "", "Investigates how time-series classifiers process local vs. global information under adversarial attack.", "", "", "", "", "", "Please review", "", "Please verify the summary and add available links."],
]
rows.extend([["", "", "", "", "", "", "", "", "", "", "", "New entry", "", ""] for _ in range(20)])
add_rows(pubs, rows)
status_dropdown(pubs, "L")
dropdown(pubs, '"Journal article,Conference paper,Workshop paper,Preprint,Book chapter,Dataset,Software,Other"', "E4:E400")
highlight_input(pubs, "A4:A400")


news = wb.create_sheet("News")
headers = ["Date (YYYY-MM-DD) — Required", "Headline — Required", "Current website description", "Your confirmed or updated description — Required", "Related URL — Optional", "Image file name — Optional", "Status", "Completed / confirmed by", "Notes"]
setup(news, "News and Updates", "Suitable items include paper acceptances, awards, events, new members, grants, talks, datasets, and software releases. Keep descriptions concise and factual.", headers, [28, 50, 68, 70, 43, 30, 20, 25, 42])
rows = [
    ["2025-11-01", "Two papers presented at CIKM 2025, Seoul", "The group presented work on calibration for Kolmogorov-Arnold Networks and medical segmentation models.", "", "", "", "Please review", "", "Only the month was known, so the date is temporarily set to the first day."],
    ["2025-09-01", "CIKM 2025 Travel Award", "Wenhao Liang received a CIKM 2025 Travel Award from the organizing committee to support participation and presentation in Seoul, Korea.", "", "", "", "Please review", "", "Only the month was known, so the date is temporarily set to the first day."],
    ["2024-08-01", "FinSen dataset released", "The FinSen financial news and sentiment dataset, spanning 197 countries, was released on GitHub.", "", "https://github.com/EagleAdelaide/FinSen_Dataset", "", "Please review", "", "Only the month was known, so the date is temporarily set to the first day."],
]
rows.extend([["", "", "", "", "", "", "New entry", "", ""] for _ in range(15)])
add_rows(news, rows)
status_dropdown(news, "G")
highlight_input(news, "D4:D300")


for ws in wb.worksheets:
    ws.sheet_properties.tabColor = BLUE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

wb.save(OUTPUT)
print(OUTPUT)
