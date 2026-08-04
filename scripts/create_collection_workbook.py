from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "DT_Lab_网站信息收集表.xlsx"

NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
PALE_YELLOW = "FFF2CC"
PALE_GREEN = "E2F0D9"
WHITE = "FFFFFF"
GREY = "666666"
THIN = Side(style="thin", color="D9E1F2")


def setup_sheet(ws, title, description, headers, widths):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = description
    ws["A2"].font = Font(size=10, color=GREY, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32

    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)
        ws.column_dimensions[cell.column_letter].width = widths[col - 1]
    ws.row_dimensions[3].height = 32
    ws.auto_filter.ref = f"A3:{ws.cell(3, len(headers)).coordinate}"


def add_rows(ws, rows):
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        ws.row_dimensions[row[0].row].height = 34


def add_status_validation(ws, column, start=4, end=200):
    validation = DataValidation(
        type="list", formula1='"待确认,已确认,需修改,新增,不公开,N/A"', allow_blank=True
    )
    ws.add_data_validation(validation)
    validation.add(f"{column}{start}:{column}{end}")


def highlight_blanks(ws, range_string):
    ws.conditional_formatting.add(
        range_string,
        FormulaRule(formula=[f'LEN({range_string.split(":")[0]})=0'], fill=PatternFill("solid", fgColor=PALE_YELLOW)),
    )


wb = Workbook()
intro = wb.active
intro.title = "填写说明"
intro.sheet_view.showGridLines = False
intro.column_dimensions["A"].width = 22
intro.column_dimensions["B"].width = 95
intro["A1"] = "DT Lab 网站信息收集表"
intro["A1"].font = Font(size=22, bold=True, color=WHITE)
intro["A1"].fill = PatternFill("solid", fgColor=NAVY)
intro["B1"].fill = PatternFill("solid", fgColor=NAVY)
intro.merge_cells("A1:B1")
intro.row_dimensions[1].height = 38

instructions = [
    ("用途", "用于收集和确认实验室网站所需信息。请按对应工作表填写，不确定的信息可以留空。"),
    ("填写人", "请在每一条信息的“填写人/确认人”栏留下姓名，便于后续核对。"),
    ("状态", "待确认＝现有信息尚未确认；已确认＝可直接公开；需修改＝请在相应字段改正；新增＝网站尚未收录；不公开＝不要放到网站；N/A＝不适用。"),
    ("黄色单元格", "建议补充或确认的内容。"),
    ("链接格式", "请填写完整网址，例如 https://example.com；邮箱只需填写 name@example.edu。"),
    ("图片", "成员照片建议为清晰正面照，JPG/PNG，尽量使用竖版或方形图。请把照片文件名写入“照片文件名”，并与表格一起提交。"),
    ("日期格式", "新闻日期统一使用 YYYY-MM-DD，例如 2026-07-31。"),
    ("论文作者", "请按论文正式署名顺序填写，用英文逗号分隔。"),
    ("隐私提醒", "只填写同意公开在实验室网站上的个人资料、邮箱和链接。"),
]
for row_idx, (label, text) in enumerate(instructions, 3):
    intro.cell(row_idx, 1, label).font = Font(bold=True, color=NAVY)
    intro.cell(row_idx, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    intro.cell(row_idx, 2, text)
    for cell in intro[row_idx]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=THIN)
    intro.row_dimensions[row_idx].height = 38

intro["A14"] = "提交前检查"
intro["A14"].font = Font(size=14, bold=True, color=WHITE)
intro["A14"].fill = PatternFill("solid", fgColor=BLUE)
intro["B14"].fill = PatternFill("solid", fgColor=BLUE)
intro.merge_cells("A14:B14")
checks = [
    "□ 实验室正式名称、简介、隶属单位和联系方式已确认",
    "□ 每位成员已确认公开资料与照片",
    "□ 研究方向描述已确认",
    "□ 论文作者顺序、发表信息和链接已确认",
    "□ 新闻日期与描述已确认",
]
for idx, text in enumerate(checks, 15):
    intro.cell(idx, 2, text)
    intro.cell(idx, 2).alignment = Alignment(wrap_text=True)
    intro.row_dimensions[idx].height = 25


site = wb.create_sheet("实验室基本信息")
site_headers = ["类别", "信息项", "网站现有内容", "请填写/修改", "状态", "填写人/确认人", "备注"]
setup_sheet(site, "实验室基本信息", "请确认实验室名称、简介、隶属关系、联系方式和网站地址。", site_headers, [16, 24, 50, 50, 14, 20, 38])
site_rows = [
    ["基本信息", "实验室正式英文名", "Data Transpose Lab", "", "待确认", "", "确认是否使用 Data Transpose Lab、DT Lab 或其他名称"],
    ["基本信息", "实验室正式中文名", "", "", "待确认", "", "如无需中文名可填 N/A"],
    ["基本信息", "英文简称", "DT Lab", "", "待确认", "", ""],
    ["基本信息", "网站英文简介", "AI, big data, and cloud-computing research at Adelaide University", "", "待确认", "", "建议 1–2 句话"],
    ["基本信息", "网站中文简介", "", "", "待确认", "", "可选"],
    ["隶属关系", "学校/大学正式名称", "Adelaide University", "", "待确认", "", "请确认学校最新正式英文名"],
    ["隶属关系", "学院/School", "School of Computer Science and Information Technology", "", "待确认", "", ""],
    ["隶属关系", "研究院/中心", "Australian Institute for Machine Learning", "", "待确认", "", "如适用"],
    ["联系信息", "实验室公共邮箱", "", "", "待确认", "", "没有可填 N/A"],
    ["联系信息", "实验室主页地址", "", "", "待确认", "", "域名或 GitHub Pages 地址"],
    ["联系信息", "完整邮寄地址", "The University of Adelaide, North Terrace, Adelaide, South Australia", "", "待确认", "", "可补充楼宇、房间和邮编"],
    ["联系信息", "办公室/房间", "", "", "待确认", "", "如适用"],
    ["联系信息", "联系电话", "", "", "待确认", "", "如需公开"],
    ["社交媒体", "GitHub 组织主页", "", "", "待确认", "", "如适用"],
    ["社交媒体", "LinkedIn/其他平台", "", "", "待确认", "", "如适用"],
]
add_rows(site, site_rows)
add_status_validation(site, "E")
highlight_blanks(site, "D4:D200")


people = wb.create_sheet("成员信息")
people_headers = [
    "成员类别", "英文姓名", "中文姓名", "职称/网站角色", "单位/隶属", "研究兴趣（英文）",
    "个人简介（英文）", "邮箱", "个人主页", "Google Scholar", "GitHub", "LinkedIn",
    "照片文件名", "所在地", "状态", "填写人/确认人", "备注",
]
setup_sheet(people, "成员信息", "每位成员一行；可在表格末尾继续新增。只填写本人同意公开的信息。", people_headers, [18, 24, 18, 28, 42, 55, 58, 30, 38, 38, 32, 32, 26, 22, 14, 20, 35])
people_rows = [
    ["Faculty", "Dr. Tony Weitong Chen", "", "Co-Director and Joint Principal Leader", "School of Computer Science and Information Technology; Australian Institute for Machine Learning", "Foundation model alignment, next-generation neural architectures, trustworthy machine learning, multimodal health data analytics, and machine unlearning.", "Senior Lecturer and ARC Externally-Funded Senior Research Fellow at Adelaide University. Eligible to supervise Master's and PhD students.", "weitong.chen@adelaide.edu.au", "https://researchers.adelaide.edu.au/profile/weitong.chen", "", "", "", "weitong-chen.jpg", "Adelaide, Australia", "待确认", "", "请确认姓名、职称和简介"],
    ["Faculty", "A/Prof. Wei Emma Zhang", "", "Co-Director and Joint Principal Leader", "School of Computer Science and Information Technology; Australian Institute for Machine Learning", "Natural language processing, text mining, information retrieval, multimodal generation and evaluation, federated learning in NLP, and Artificial Intelligence of Things.", "Associate Professor and ARC Externally-Funded Senior Research Fellow at Adelaide University. Eligible to supervise Master's and PhD students.", "wei.e.zhang@adelaide.edu.au", "https://researchers.adelaide.edu.au/profile/wei.e.zhang", "", "", "", "wei-emma-zhang.jpg", "Adelaide, Australia", "待确认", "", "请确认姓名、职称和简介"],
    ["PhD Student", "Wenhao Liang", "", "PhD Candidate", "Australian Institute for Machine Learning; SA Pathology; The University of Adelaide", "Trust, safety, security, privacy, uncertainty estimation, model calibration, and model optimization.", "", "", "http://eaglewhliang.github.io/", "https://scholar.google.com/citations?user=3f4N-oQAAAAJ&hl=en", "https://github.com/EagleAdelaide", "", "wenhao-liang.jpg", "Adelaide, Australia", "待确认", "", "请补充邮箱和简介（如公开）"],
]
for _ in range(12):
    people_rows.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "新增", "", ""])
add_rows(people, people_rows)
add_status_validation(people, "O")
member_type = DataValidation(type="list", formula1='"Faculty,Postdoc,PhD Student,Master Student,Undergraduate,Visitor,Alumni,Staff"', allow_blank=True)
people.add_data_validation(member_type)
member_type.add("A4:A200")
highlight_blanks(people, "B4:B200")


research = wb.create_sheet("研究方向")
research_headers = ["排序", "研究方向英文名称", "研究方向中文名称", "网站现有英文描述", "请填写/修改英文描述", "代表关键词", "是否首页展示", "状态", "填写人/确认人", "备注"]
setup_sheet(research, "研究方向", "每个研究方向一行；描述建议 25–60 个英文词，关键词用英文逗号分隔。", research_headers, [10, 30, 25, 65, 65, 45, 18, 14, 20, 35])
research_rows = [
    [1, "Trustworthy AI", "可信人工智能", "Model calibration, uncertainty estimation, fairness, adversarial robustness, privacy, interpretability, and explainability.", "", "", "是", "待确认", "", ""],
    [2, "AI and Data Mining", "人工智能与数据挖掘", "NLP, text mining, causal reasoning, multimodal information fusion, and knowledge discovery from imperfect data.", "", "", "是", "待确认", "", ""],
    [3, "Model Optimization", "模型优化", "Efficient training and inference, hyperparameter and optimization methods, transfer learning, meta-learning, and hardware-aware AI.", "", "", "是", "待确认", "", ""],
    [4, "Big Data and Cloud Computing", "大数据与云计算", "Big data, distributed systems, and cloud computing.", "", "", "待确认", "待确认", "", "确认是否需要首页重点展示"],
]
for idx in range(5, 9):
    research_rows.append([idx, "", "", "", "", "", "", "新增", "", ""])
add_rows(research, research_rows)
add_status_validation(research, "H")
yes_no = DataValidation(type="list", formula1='"是,否,待确认"', allow_blank=True)
research.add_data_validation(yes_no)
yes_no.add("G4:G200")
highlight_blanks(research, "E4:E200")


pubs = wb.create_sheet("论文成果")
pub_headers = ["论文标题", "作者（按正式顺序）", "会议/期刊", "年份", "论文类型", "简短摘要/亮点", "DOI链接", "arXiv链接", "PDF链接", "代码链接", "项目主页", "状态", "填写人/确认人", "备注"]
setup_sheet(pubs, "论文成果", "每篇论文一行；作者请按正式发表顺序填写，链接请填写完整网址。", pub_headers, [55, 65, 35, 10, 18, 60, 35, 35, 35, 35, 35, 14, 20, 35])
pub_rows = [
    ["Calibrating on Kolmogorov-Arnold Networks", "Wenhao Liang, Wei Emma Zhang, Lin Yue, Miao Xu, Olaf Maennel, Weitong Chen", "CIKM 2025, Seoul", 2025, "", "", "", "https://arxiv.org/abs/2503.01195", "", "", "", "待确认", "", "确认书目信息并补充链接"],
    ["We Care Each Pixel: Calibrating on Medical Segmentation Models through Signed Distance", "Wenhao Liang, Wei Emma Zhang, Lin Yue, Miao Xu, Olaf Maennel, Weitong Chen", "CIKM 2025, Seoul", 2025, "", "", "", "https://arxiv.org/abs/2503.05107", "", "", "", "待确认", "", "确认书目信息并补充链接"],
    ["TraffiX-MoE: A Traffic-Aware Neural VRP Solver", "Wenhao Liang, Lin Yue, Wei Emma Zhang, Joy Rathjen, Peter O'Loughlin, Weitong Chen", "ADMA 2025, Kyoto, Industry Track", 2025, "", "", "", "", "", "", "", "待确认", "", "补充论文链接"],
    ["Enhancing Financial Market Predictions: Causality-Driven Feature Selection", "Wenhao Liang, Zhengyang Li, Weitong Chen", "ADMA 2024", 2024, "", "Integrates financial news and stock-market data across 197 countries with LSTM models to improve market-prediction accuracy.", "", "", "", "", "", "待确认", "", "确认摘要并补充链接"],
    ["Correlation Analysis of Adversarial Attack in Time Series Classification", "Zhengyang Li, Wenhao Liang, Chang Dong, Weitong Chen, Dong Huang", "ADMA 2024", 2024, "", "Investigates how time-series classifiers process local vs. global information under adversarial attack.", "", "", "", "", "", "待确认", "", "确认摘要并补充链接"],
]
for _ in range(15):
    pub_rows.append(["", "", "", "", "", "", "", "", "", "", "", "新增", "", ""])
add_rows(pubs, pub_rows)
add_status_validation(pubs, "L")
pub_type = DataValidation(type="list", formula1='"Journal,Conference,Workshop,Preprint,Book Chapter,Dataset,Other"', allow_blank=True)
pubs.add_data_validation(pub_type)
pub_type.add("E4:E300")
highlight_blanks(pubs, "A4:A300")


news = wb.create_sheet("新闻动态")
news_headers = ["日期（YYYY-MM-DD）", "英文标题", "中文标题（可选）", "网站现有英文描述", "请填写/修改英文描述", "相关链接", "相关图片文件名", "状态", "填写人/确认人", "备注"]
setup_sheet(news, "新闻动态", "每条新闻一行；可收集论文接收、获奖、活动、成员加入、项目发布等。", news_headers, [22, 45, 36, 65, 65, 40, 28, 14, 20, 38])
news_rows = [
    ["2025-11-01", "Two papers presented at CIKM 2025, Seoul", "", "The group presented work on calibration for Kolmogorov-Arnold Networks and medical segmentation models.", "", "", "", "待确认", "", "原始信息只有月份，日期暂设为 01"],
    ["2025-09-01", "CIKM 2025 Travel Award", "", "Wenhao Liang received a CIKM 2025 Travel Award from the organizing committee to support participation and presentation in Seoul, Korea.", "", "", "", "待确认", "", "原始信息只有月份，日期暂设为 01"],
    ["2024-08-01", "FinSen dataset released", "", "The FinSen financial news and sentiment dataset, spanning 197 countries, was released on GitHub.", "", "https://github.com/EagleAdelaide/FinSen_Dataset", "", "待确认", "", "原始信息只有月份，日期暂设为 01"],
]
for _ in range(12):
    news_rows.append(["", "", "", "", "", "", "", "新增", "", ""])
add_rows(news, news_rows)
add_status_validation(news, "H")
highlight_blanks(news, "B4:B200")


for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.tabColor = BLUE
    ws.auto_filter.ref = ws.auto_filter.ref if ws.auto_filter.ref else None

wb.save(OUTPUT)
print(OUTPUT)
